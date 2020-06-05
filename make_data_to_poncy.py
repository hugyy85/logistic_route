from pyroutelib3 import Router
import geopy.distance
from openpyxl import load_workbook

import requests
import os

from models import Address, DistanceBetweenAddress

import time


YANDEX_API_KEY = ''


class Route:

    def __init__(self, start_point, xlsx_file_path, yandex_api_key, number_routes=2, end_point=None, city='Новосибирск',
                 calculate_method='car'):
        self.start_point = start_point
        self.workbook = load_workbook(xlsx_file_path)
        self.yandex_api_key = yandex_api_key
        # количество маршрутов на которое делить общий маршрут, чтобы получилось равномерно(например количество машин)
        self.number_routes = number_routes
        self.end_point = end_point
        self.city = city
        self.calculate_method = calculate_method

    def get_coordinates_from_yandex(self, address):
        url = 'https://geocode-maps.yandex.ru/1.x/'
        params = {
            'apikey': os.getenv('YANDEX_API_KEY', self.yandex_api_key),
            'geocode': address,
            'format': 'json'
        }
        response = requests.get(url, params=params).json()
        try:
            lat, lng = response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['Point'][
                'pos'].split(' ')
        except KeyError:
            raise Exception(f'Error, request - {response}')

        return float(lat), float(lng)

    def get_distance_between_two_addresses(self, first_address, second_address):
        # check address in db
        query_first = Address.select().where(Address.name == first_address)
        query_second = Address.select().where(Address.name == second_address)

        if query_first.exists() and query_second.exists():
            query_distance = DistanceBetweenAddress.select().where(
                DistanceBetweenAddress.address_id == query_first.get().id,
                DistanceBetweenAddress.next_address_id == query_second.get().id)
            if query_distance.exists():
                return query_distance.get().distance

        router = Router(self.calculate_method)

        # Initialise it
        coord_s_t = time.time()
        first_lat, first_lng = (query_first.get().lat, query_first.get().lng) if query_first.exists() else \
            self.get_coordinates_from_yandex(first_address)

        second_lat, second_lng = (query_second.get().lat, query_second.get().lng) if query_second.exists() else \
            self.get_coordinates_from_yandex(second_address)

        coord_e_t = time.time()

        start = router.findNode(first_lng, first_lat)  # Find start and end nodes
        end = router.findNode(second_lng, second_lat)
        rout_s_t = time.time()
        status, route = router.doRoute(start, end)  # Find the route - a list of OSM nodes
        route_e_t = time.time()

        if status == 'success':
            routeLatLons = list(map(router.nodeLatLon, route))  # Get actual route coordinates
            total_distance = 0
            # calculate total distance from route coordinates
            for index in range(1, len(routeLatLons)):
                total_distance += geopy.distance.vincenty(routeLatLons[index - 1], routeLatLons[index]).km
        else:
            total_distance = 0
            # это случается, когда 2 точки с одинаковым адресом, надо перепроверить
            print(f'{route}')
        first = query_first.get().id if query_first.exists() else Address.create(name=first_address,
                                                                                 lat=first_lat, lng=first_lng)
        second = query_second.get().id if query_second.exists() else Address.create(name=second_address,
                                                                                    lat=second_lat, lng=second_lng)
        DistanceBetweenAddress.bulk_create([
            DistanceBetweenAddress(address_id=first, next_address_id=second, distance=total_distance),
            DistanceBetweenAddress(address_id=second, next_address_id=first, distance=total_distance)
        ])

        print(f'COORDINATES TIME = {coord_e_t - coord_s_t} sec')
        print(f'ROUTE TIME = {route_e_t - rout_s_t} sec')
        print(total_distance, 'Total distance ==============================')
        return total_distance

    def calculate_route(self, start_point, addresses):
        """
        Алгоритм 1)
        Осуществляем поиск всех расстояний от точки начала, сортируем по возрастанию, делим точки на количество автомобилей
        Повторяем, но с ближайшеми 5 точками, потом завести тест, в котором будет проверятся сильно ли менялся маршрут после
        пересчета

        :param addresses:
        :param city: str Город где строится маршрут
        :param end_point: str Точка конца маршрута, если None то равно точке начала
        :param start_point: str Точка начала маршрута
        :param file_with_points_route: str путь до файла откуда вытаскивать данные(*.xlsx)
        :return:
        """

        map_distances = {}
        for address in addresses:
            if address:
                distance = self.get_distance_between_two_addresses(f'{self.city} {start_point}', f'{self.city} {address}')
                map_distances[address] = distance
            else:
                # todo Придумать куда записывать точки без адресов
                #
                continue
        sorted_by_value = sorted(map_distances.items(), key=lambda kv: kv[1])
        return sorted_by_value[0] if sorted_by_value else ("END_ROUTE", 0)

    def make_route(self):
        end_point = self.end_point or self.start_point
        start_t = time.time()
        sheet = self.workbook.get_sheet_by_name('TDSheet')
        addresses = [self.start_point]
        for row in range(2, sheet.max_row + 1):
            addresses.append(sheet[f'E{row}'].value)

        result = []
        total_distance = 0
        for _ in addresses:
            # удаляем начальную точку, чтобы в следующей точке сделать поиск по найденой
            # алгормитм: берем точку начала - ищем расстояние до каждой точки от точки начала, находим наименьшее,
            # и точку с наименьшим расстоянием принимает за точку начала, удаляем эту точку из адрессов чтобы не искать
            # расстояние до этой точки(так как мы находимся в ней). И повторяем до тех пор, пока не адреса не закончатся
            # если точки были проверены ранее, то берем их из БД, не просчитывая маршрут

            addresses.remove(self.start_point)
            self.start_point, distance = self.calculate_route(self.start_point, addresses)
            if self.start_point == 'END_ROUTE':
                last_address = result[len(result) - 1][0]
                distance = self.get_distance_between_two_addresses(last_address, f'{self.city} {end_point}')
                result.append((end_point, distance))
                # todo обработать конец маршрута и возвращение на базу. Идея создать базу данных где будет храниться каждая точка
                # организации и расстояние от этой точки до всех других точек организации!!!!

                break
            result.append((self.start_point, distance))
            total_distance += distance
        print(result, total_distance)
        end_t = time.time()

        print(f'{end_t - start_t} seconds all script')

    def add_manual_correct(self):
        pass


route = Route('Выставочная 19', 'Реестр 02 06.xlsx', YANDEX_API_KEY)
route.make_route()
